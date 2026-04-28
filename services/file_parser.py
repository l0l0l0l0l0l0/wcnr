# -*- coding: utf-8 -*-
"""
Excel/CSV 文件解析与校验
"""

from __future__ import annotations

import csv
import io
import re
from typing import Optional

ID_CARD_RE = re.compile(r"^\d{17}[\dXx]$")

POPULATION_COLUMN_MAP = {
    "姓名": "name",
    "身份证号": "id_card_number",
    "身份证号码": "id_card_number",
    "公民身份证号码": "id_card_number",
    "性别": "gender",
    "年龄": "age",
    "住址": "address",
    "居住地详址": "address",
    "地址": "address",
    "联系方式": "contact",
    "联系电话": "contact",
    "电话": "contact",
    "手机": "contact",
}

CASE_COLUMN_MAP = {
    "案件编号": "case_number",
    "案件名称": "case_name",
    "案件类型": "case_type",
    "案发时间": "incident_time_str",
    "案发地点": "incident_location",
    "涉案人员": "involved_persons",
}


def parse_upload_file(file_content: bytes, filename: str, column_map: dict) -> list[dict]:
    """解析上传的 Excel 或 CSV 文件，返回行字典列表。"""
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

    if ext in ("xlsx", "xls"):
        return _parse_excel(file_content, column_map)
    elif ext == "csv":
        return _parse_csv(file_content, column_map)
    else:
        raise ValueError(f"不支持的文件格式: {ext}，请上传 xlsx/xls/csv 文件")


def _parse_excel(content: bytes, column_map: dict) -> list[dict]:
    from openpyxl import load_workbook

    wb = load_workbook(filename=io.BytesIO(content), read_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)

    header = next(rows_iter, None)
    if header is None:
        wb.close()
        return []

    col_indices = _map_columns(header, column_map)
    result = []
    for row_idx, row in enumerate(rows_iter, start=2):
        mapped = _map_row(row, col_indices)
        mapped["_row_number"] = row_idx
        result.append(mapped)

    wb.close()
    return result


def _parse_csv(content: bytes, column_map: dict) -> list[dict]:
    text = content.decode("utf-8-sig")
    reader = csv.reader(io.StringIO(text))
    header = next(reader, None)
    if header is None:
        return []

    col_indices = _map_columns(header, column_map)
    result = []
    for row_idx, row in enumerate(reader, start=2):
        padded = list(row) + [None] * (len(col_indices) - len(row) + 1)
        mapped = _map_row(padded, col_indices)
        mapped["_row_number"] = row_idx
        result.append(mapped)

    return result


def _map_columns(header: tuple | list, column_map: dict) -> dict:
    """将表头映射为 {字段名: 列索引}。"""
    result = {}
    for idx, col_name in enumerate(header):
        col_name = str(col_name).strip() if col_name else ""
        if col_name in column_map:
            result[column_map[col_name]] = idx
    return result


def _map_row(row: tuple | list, col_indices: dict) -> dict:
    """按列映射提取行数据。"""
    result = {}
    for field_name, col_idx in col_indices.items():
        val = row[col_idx] if col_idx < len(row) else None
        if val is not None:
            val = str(val).strip()
            if val == "":
                val = None
        result[field_name] = val
    return result


def validate_population_row(row: dict) -> tuple[bool, Optional[str]]:
    """校验人口数据行。返回 (is_valid, error_message)。"""
    id_card = row.get("id_card_number")
    if not id_card:
        return False, "身份证号不能为空"
    if not ID_CARD_RE.match(id_card):
        return False, f"身份证号格式不正确: {id_card}"
    if not row.get("name"):
        return False, "姓名不能为空"
    return True, None


def validate_case_row(row: dict) -> tuple[bool, Optional[str]]:
    """校验案件数据行。返回 (is_valid, error_message)。"""
    case_number = row.get("case_number")
    if not case_number:
        return False, "案件编号不能为空"
    if not row.get("case_name"):
        return False, "案件名称不能为空"
    return True, None


def parse_involved_persons(text: Optional[str]) -> list[dict]:
    """
    解析 '涉案人员' 单元格文本为结构化人员列表。
    支持格式: '姓名/身份证号' 或单独身份证号，以逗号/分号/换行分隔。
    """
    if not text:
        return []

    result = []
    for part in re.split(r"[,;，；\n]", text):
        part = part.strip()
        if not part:
            continue

        if "/" in part:
            segments = part.split("/", 1)
            name = segments[0].strip()
            id_card = segments[1].strip() if len(segments) > 1 else None
        else:
            # 纯数字则视为身份证号
            if re.match(r"^\d{17}[\dXx]$", part):
                name = None
                id_card = part
            else:
                name = part
                id_card = None

        if name or id_card:
            result.append({"person_name": name, "id_card_number": id_card})

    return result
